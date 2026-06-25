from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import os

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

QUESTIONS = [
    {"section": "ВХОДНАЯ ГРУППА", "section_fill": "FFDDD9C3", "text": "Входная группа чистая", "importance": 2},
    {"section": None, "text": "Вывеска горит", "importance": 1},
    {"section": None, "text": "Сотрудники приветливо и радушно встречают гостей, предлагают стол и помочь с размещением - подставка под сумку/меню, нет потерянных гостей, которые не понимают, куда им идти", "importance": 2},
    {"section": None, "text": "Вся зона входа и в полном порядке - гардероб/ вешалки/циновки на столах и тд, нет ничего лишнего - первое впечатление после входа в ресторан - полный порядок в деталях", "importance": 2},
    {"section": None, "text": "Сотрудники работают в команде, чувствуется дружелюбная атмосфера, на столах у гостей нет грязной посуды", "importance": 1},
    {"section": None, "text": "Администратор в зале есть, помогает с сервисом, включен в рабочий процесс", "importance": 2},
    {"section": "ПОДРАЗДЕЛЕНИЯ РЕСТОРАНА И ПЕРВОЕ ВПЕЧАТЛЕНИЕ", "section_fill": "FFFBD4B4", "text": "Меню и все печатные материалы в хорошем состоянии, чистые, без заломов, актуальные, все страницы в наличии", "importance": 3},
    {"section": None, "text": "Вентиляция работает бесшумно, посторонних запахов и запахов еды нет, задымленность отсутствует", "importance": 2},
    {"section": None, "text": "Громкость музыки комфортная - не громко, но и не тихо, играет подходящий репертуар, освещение соответствует времени суток, над столами включен свет, отсутствуют перегоревшие лампы", "importance": 2},
    {"section": None, "text": "Флористика во всем ресторане в хорошем состоянии - большие вазы, вазы в уборных, вазы на столах, цветы на входных зонах", "importance": 1},
    {"section": None, "text": "Форма всех сотрудников зала чистая, выглаженная, обувь без повреждений, внешний вид в полном порядке, волосы собраны, сотрудник выглядит соответствующе", "importance": 2},
    {"section": None, "text": "Чистота везде: стены, мебель, вытяжки, решетки на стенах, поверхности столов, станции официанта, холодильники, все полки и рабочие поверхности; на рабочих поверхностях также отсутствуют личные вещи", "importance": 2},
    {"section": None, "text": "Санузел чистый, зеркала без подтеков, столешницы и сантехника чистые", "importance": 2},
    {"section": None, "text": "Детали интерьера, мебель - все в полном порядке и в хорошем состоянии, столы не шатаются, жвачек на обратной стороне столов нет, диваны и кресла чистые без волос и пыли", "importance": 3},
    {"section": None, "text": "Сотрудники доставки выглядят опрятно, находятся на рабочем месте", "importance": 1},
    {"section": None, "text": "Сервировка расставлена на столах/баре, вся посуда чистая, без сколов, лишняя сервировка снимается", "importance": 1},
    {"section": "ОСОБЕННОСТИ РЕСТОРАНА", "section_fill": "FFB8CCE4", "text": "Сотрудники не сидят на стульях для гостей, все распределены четко по позициям", "importance": 2},
    {"section": None, "text": "Шторы во всем ресторане чистые и открыты/закрыты в зависимости от времени суток", "importance": 3},
    {"section": None, "text": "Сотрудники улыбаются гостям, внимательно контролируют посадку", "importance": 1},
    {"section": "СОБЛЮДЕНИЕ СТАНДАРТОВ СЕРВИСА", "section_fill": "FFFFF2CC", "text": "Сотрудники подают меню открытой рукой, в открытом виде и прямо в руки, дополнительное меню размещают на столе и комментируют свои действия", "importance": 2},
    {"section": None, "text": "Сотрудники предлагают заказать напитки на этапе подачи меню", "importance": 2},
    {"section": None, "text": "Сотрудники хорошо знают меню", "importance": 3},
    {"section": None, "text": "Сотрудники хорошо работают по доп продажам", "importance": 2},
    {"section": None, "text": "Сотрудники уточняют последовательность подачи блюд и повторяют заказ", "importance": 2},
    {"section": None, "text": "Все блюда и напитки, которые выносятся гостю соответствуют стандарту", "importance": 2},
    {"section": None, "text": "Сотрудники запрашивают обратную связь", "importance": 2},
    {"section": None, "text": "Сотрудники своевременно зачищают столы", "importance": 1},
    {"section": None, "text": "Доставка собирается корректно, в термо-пакетах, брендированных пакетах и коробках", "importance": 1},
    {"section": None, "text": "Сотрудники быстро приносят счёт и своевременно рассчитывают гостей", "importance": 1},
    {"section": "КУХНЯ", "section_fill": "FFD6E3BC", "text": "Блюда на раздаче не стоят дольше 1 минуты", "importance": 1},
    {"section": None, "text": "Чистота всех поверхностей и пола в зоне кухни, порядок на стеллажах и раздаче, нет лишних предметов и личных вещей", "importance": 2},
    {"section": None, "text": "Все блюда по вкусу и виду подачи соответствуют стандарту", "importance": 1},
    {"section": None, "text": "Не нарушены условия хранения продукции, везде присутствует маркировка", "importance": 1},
    {"section": None, "text": "Все сотрудники кухни выглядят опрятно, волосы убраны", "importance": 2},
    {"section": "БАР", "section_fill": "FFDAEEF3", "text": "Чистота всех поверхностей, стен и пола в зоне бара ресторана", "importance": 2},
    {"section": None, "text": "Все позиции бара в наличии/ вина/ крепкий алкоголь/ лимонады/ пунши и тд", "importance": 3},
    {"section": None, "text": "Все стекло, барная посуда и инвентарь чистые и натертые, без сколов", "importance": 2},
    {"section": None, "text": "Все сотрудники выглядят опрятно, волосы убраны", "importance": 2},
    {"section": None, "text": "Напитки на раздачах не стоят дольше 1 минуты", "importance": 2},
    {"section": "СОТРУДНИКИ", "section_fill": "FFCCC0D9", "text": "Сотрудники клининговой службы выглядит опрятно, одеты в униформу, обувь закрытая; Передвигаются по залу незаметно и не мешают нахождению гостей в ресторане", "importance": 2},
    {"section": None, "text": "Все сотрудники общаются в естественной, дружелюбной и заинтересованной манере друг с другом и с гостями, проявляют инициативу при общении с гостем, запрашивают обратную связь, поддерживают с ним зрительный контакт, не чувствуется негатива в коллективе", "importance": 2},
    {"section": None, "text": "Отсутствует запах изо рта, запах пота, сигарет, резкий парфюм, волосы у сотрудников убранные и чистые, аккуратный маникюр", "importance": 2},
    {"section": None, "text": "Сотрудники не пользуются мобильными телефонами в зале, не расчесывают волосы, не трогают лицо/уши, не жуют жвачку", "importance": 2},
    {"section": None, "text": "В речи всех сотрудников отсутствуют слова-паразиты, мат, жаргонизмы, профессиональный сленг", "importance": 2},
    {"section": "Общее впечатление", "section_fill": "FFD6E3BC", "text": "Общее впечатление от посещения ресторана положительное", "importance": 1},
]

def make_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def make_border(left=None, right=None, top=None, bottom=None):
    def side(style):
        return Side(border_style=style, color='FF000000') if style else Side()
    return Border(left=side(left), right=side(right), top=side(top), bottom=side(bottom))

def make_font(bold=False, color='FF000000', size=11):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def make_align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def generate_excel(data):
    restaurant = data.get('restaurant', '')
    audit_date = data.get('date', '')
    auditor = data.get('auditor', '')
    manager = data.get('manager', '')
    answers = data.get('answers', [])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Бланк Тайный Гость'

    ws.column_dimensions['A'].width = 5.16
    ws.column_dimensions['B'].width = 4.0
    ws.column_dimensions['C'].width = 34.16
    ws.column_dimensions['D'].width = 41.5
    ws.column_dimensions['E'].width = 19.16
    ws.column_dimensions['F'].width = 71.33
    ws.column_dimensions['G'].width = 33.83
    ws.column_dimensions['H'].width = 27.33
    ws.column_dimensions['I'].width = 8.83

    fill_pink = make_fill('FFE5B8B7')
    fill_blue = make_fill('FFDBE5F1')
    fill_hdr  = make_fill('FFB2A1C7')
    fill_red  = make_fill('FFFF0000')

    ws.merge_cells('A1:C2'); c=ws['A1']; c.value='Ресторан:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('D1:D2'); c=ws['D1']; c.value=restaurant; c.fill=fill_blue; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('E1:E2'); c=ws['E1']; c.value='Дата посещения:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('F1:F2'); c=ws['F1']; c.value=audit_date; c.fill=fill_blue; c.font=make_font(); c.alignment=make_align()

    ws.merge_cells('A3:C4'); c=ws['A3']; c.value='Имя проверяющего:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('D3:D4'); c=ws['D3']; c.value=auditor; c.fill=fill_blue; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('E3:E4'); c=ws['E3']; c.value='Время прихода:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('F3:F4'); ws['F3'].fill=fill_blue

    ws.merge_cells('A5:C6'); c=ws['A5']; c.value='Имя менеджера на смене:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('D5:D6'); c=ws['D5']; c.value=manager; c.fill=fill_blue; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('E5:E6'); c=ws['E5']; c.value='Время ухода:'; c.fill=fill_pink; c.font=make_font(); c.alignment=make_align()
    ws.merge_cells('F5:F6'); ws['F5'].fill=fill_blue
    ws['H6'].value='Система оценки:'; ws['H6'].font=make_font(bold=True)

    ws.row_dimensions[7].height = 16
    for coord, val, bl, br in [
        ('A7','Важность','medium','thin'), ('B7','№','thin','thin'),
        ('C7','Раздел','thin','thin'), ('D7','Подпункт','thin','thin'),
        ('E7','Оценка','thin','thin'), ('F7','Комментарий','thin','medium')
    ]:
        c=ws[coord]; c.value=val; c.fill=fill_hdr; c.font=make_font()
        c.alignment=make_align(); c.border=make_border(left=bl,right=br,top='medium',bottom='thin')

    ws.merge_cells('H7:I7')
    ws['H7'].value='0 - не выполнено'
    ws['H8'].value='1 - выполнено'
    ws['H9'].value='N - не применимо в данной ситуации'

    section_spans = {}
    cur_sec = cur_start = cur_fill = None
    ROW_START = 8
    for i, q in enumerate(QUESTIONS):
        row = ROW_START + i
        if q.get('section'):
            if cur_sec: section_spans[cur_sec] = (cur_start, row-1, cur_fill)
            cur_sec, cur_start, cur_fill = q['section'], row, q.get('section_fill','FFE0E0E0')
    if cur_sec: section_spans[cur_sec] = (cur_start, ROW_START+len(QUESTIONS)-1, cur_fill)

    section_written = set()
    for i, q in enumerate(QUESTIONS):
        row = ROW_START + i
        a = answers[i] if i < len(answers) else {'score': None, 'comment': ''}
        score = a.get('score')
        comment = a.get('comment', '') or ''

        sec_name = None
        for j in range(i+1):
            if QUESTIONS[j].get('section'): sec_name = QUESTIONS[j]['section']
        sec_fill_color = section_spans.get(sec_name, (None,None,'FFE0E0E0'))[2]

        c=ws.cell(row=row,column=1); c.value=q['importance']; c.font=make_font(color='FFA5A5A5')
        c.alignment=make_align(); c.border=make_border(left='medium',right='thin',top='thin',bottom='thin')

        c=ws.cell(row=row,column=2); c.value=i+1; c.font=make_font()
        c.alignment=make_align(); c.border=make_border(left='thin',right='thin',top='thin',bottom='thin')

        c=ws.cell(row=row,column=3)
        if q.get('section') and q['section'] not in section_written:
            c.value=q['section']; c.font=make_font(bold=True); section_written.add(q['section'])
        c.fill=make_fill(sec_fill_color); c.alignment=make_align()
        c.border=make_border(left='thin',right='thin',top='thin',bottom='thin')

        c=ws.cell(row=row,column=4); c.value=q['text']; c.font=make_font(size=10)
        c.alignment=make_align('left','center',True)
        c.border=make_border(left='thin',right='thin',top='thin',bottom='thin')

        c=ws.cell(row=row,column=5)
        c.value='N' if score=='N' else (1 if score==1 else (0 if score==0 else None))
        c.font=make_font(); c.alignment=make_align('right','center',False)
        c.border=make_border(left='thin',right='thin',top='thin',bottom='thin')

        c=ws.cell(row=row,column=6); c.value=comment; c.font=make_font(size=10)
        c.alignment=make_align('left','center',True)
        c.border=make_border(left='thin',right='medium',top='thin',bottom='thin')

    for sec,(start,end,fc) in section_spans.items():
        if start < end: ws.merge_cells(f'C{start}:C{end}')
        for r in range(start,end+1):
            c=ws.cell(row=r,column=3); c.fill=make_fill(fc)
            c.font=make_font(bold=True); c.alignment=make_align()

    last_row = ROW_START + len(QUESTIONS) - 1
    sum_row = last_row + 5

    ws[f'C{sum_row}'].value='Число ответов ДА'
    ws[f'D{sum_row}'].value=f'=COUNTIF(E{ROW_START}:E{last_row},"=1")'
    ws[f'C{sum_row+1}'].value='Число ответов НЕТ'
    ws[f'D{sum_row+1}'].value=f'=COUNTIF(E{ROW_START}:E{last_row},"=0")'
    ws[f'C{sum_row+2}'].value='Число ответов N'
    ws[f'D{sum_row+2}'].value=f'=COUNTIF(E{ROW_START}:E{last_row},"N")'

    score_row = sum_row + 3
    c=ws[f'C{score_row}']; c.value='ОБЩАЯ ОЦЕНКА АУДИТА'; c.fill=fill_red
    c.font=make_font(bold=True,color='FFFFFFFF')
    c=ws[f'D{score_row}']
    c.value=f'=IFERROR(COUNTIF(E{ROW_START}:E{last_row},"=1")/(COUNTIF(E{ROW_START}:E{last_row},"=1")+COUNTIF(E{ROW_START}:E{last_row},"=0")),0)'
    c.font=make_font(bold=True,color='FFFF0000')
    c.number_format='0%'
    ws[f'J{score_row+3}'].value='ТУТ НАХОДЯТСЯ ФОРМУЛЫ, НЕ СТИРАТЬ!'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/')
def index():
    return send_file('static/index.html')

@app.route('/generate', methods=['POST'])
def generate():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    try:
        excel_file = generate_excel(data)
        restaurant = data.get('restaurant', 'Аудит')
        date = data.get('date', '')
        filename = f"{restaurant}_{date}.xlsx"
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
